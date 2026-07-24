#!/usr/bin/env python3
"""Self-contained builder for the P1 & P2 Incidents Dashboard (no external repo/module deps)."""

import base64
import datetime
import io
import json
import re
import time

import requests
import openpyxl

SPREADSHEET_ID = "15tIGNbDVZ7bHnukOztbZsztUt5JmymeDMDwcyej6J-s"
MIN_DATE = datetime.date(2025, 1, 1)
import os
HERE = os.path.dirname(os.path.abspath(__file__))
TEMPLATE_PATH = os.path.join(HERE, "dashboard_template.html")
OUTPUT_PATH = os.path.join(HERE, "index.html")

RETRY_ATTEMPTS = 3
RETRY_DELAY = 20

MONTH_ABBR = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
              7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
MONTH_FULL = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June',
              7: 'July', 8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}


def with_retry(fn, label):
    last_err = None
    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            return fn()
        except Exception as exc:
            last_err = exc
            if attempt < RETRY_ATTEMPTS:
                print(f"  [{label}] attempt {attempt} failed: {exc} — retrying in {RETRY_DELAY}s …")
                time.sleep(RETRY_DELAY)
    raise last_err


def download_workbook():
    def _download():
        url = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
        resp = requests.get(url, timeout=30)
        resp.raise_for_status()
        return openpyxl.load_workbook(io.BytesIO(resp.content), data_only=True)
    return with_retry(_download, "download")


_MONTH_FMTS = ("%B %Y", "%b %Y")


def _parse_tab_date(name):
    for fmt in _MONTH_FMTS:
        try:
            return datetime.datetime.strptime(name.strip(), fmt).date()
        except ValueError:
            pass
    return None


def get_monthly_sheets(wb, limit=None):
    result = []
    for name in wb.sheetnames:
        d = _parse_tab_date(name)
        if d:
            result.append((d, wb[name]))
    result.sort(key=lambda x: x[0], reverse=True)
    return result[:limit] if limit else result


def _norm(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())


_COL_KEYS = {
    'cause_type':    ['incidentcause'],
    'cause':         ['causerootcause', 'rootcause'],
    'incident_id':   ['incidentid'],
    'country':       ['country'],
    'product':       ['product'],
    'date':          ['date'],
    'priority':      ['priority'],
    'duration':      ['duration', 'mttr'],
    'title':         ['systemarea', 'system'],
    'impact':        ['impactsummary', 'impact'],
    'resolution_dt': ['resolutiondatetime', 'resolutiondate'],
    'resolution':    ['resolution'],
    'status':        ['statusclosed', 'status'],
    'jira_url':      ['jiraticket', 'jira'],
    'postmortem_url': ['postmortemticket', 'postmortem'],
}


def build_col_map(header_row):
    normed = [(_norm(h), i) for i, h in enumerate(header_row)]
    col_map = {}
    for field, patterns in _COL_KEYS.items():
        for pat in patterns:
            for nh, idx in normed:
                if pat in nh:
                    if field not in col_map:
                        col_map[field] = idx
    return col_map


def get_field(row, col_map, field, default=""):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return default
    return str(row[idx]).strip()


_DATE_FMTS = [
    "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M",
    "%m/%d/%Y %H:%M:%S", "%m/%d/%Y %H:%M", "%Y/%m/%d", "%Y-%m-%d",
]

_FUZZY_DATE_FMTS = [
    "%d %b %Y %I:%M %p", "%d %B %Y %I:%M %p",
    "%d %b %Y %I %p", "%d %B %Y %I %p",
    "%d %b %Y %H:%M:%S", "%d %B %Y %H:%M:%S",
    "%d %b %Y %H:%M", "%d %B %Y %H:%M",
    "%d %b %Y", "%d %B %Y",
    "%B %d, %Y %I:%M %p", "%b %d, %Y %I:%M %p",
    "%B %d, %Y", "%b %d, %Y",
]

_ORDINAL_RE = re.compile(r'(\d+)(st|nd|rd|th)\b', re.IGNORECASE)
_TZ_RE = re.compile(r'\b(EAT|GMT|UTC|EST|CAT|WAT)\b', re.IGNORECASE)
_AMPM_SPACE_RE = re.compile(r'(\d)(AM|PM)', re.IGNORECASE)
_JIRA_URL_RE = re.compile(r'(https?://\S*/browse/([A-Z][A-Z0-9]*-\d+))')


def parse_jira_links(raw):
    if not raw:
        return []
    matches = _JIRA_URL_RE.findall(raw)
    if matches:
        seen = set()
        links = []
        for url, jid in matches:
            if jid in seen:
                continue
            seen.add(jid)
            links.append({'id': jid, 'url': url})
        return links
    return [{'id': raw, 'url': None}]


def _clean_fuzzy_date(s):
    s = _ORDINAL_RE.sub(r'\1', s)
    s = _TZ_RE.sub('', s)
    s = _AMPM_SPACE_RE.sub(r'\1 \2', s)
    s = s.replace(' at ', ' ')
    return re.sub(r'\s+', ' ', s).strip()


def _try_parse_dt(s):
    s = str(s).strip()
    for fmt in _DATE_FMTS:
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            pass
    cleaned = _clean_fuzzy_date(s)
    for fmt in _FUZZY_DATE_FMTS:
        try:
            return datetime.datetime.strptime(cleaned, fmt)
        except ValueError:
            pass
    return None


def fmt_incident_date(raw):
    dt = _try_parse_dt(raw)
    return dt.strftime("%-d %b %Y, %I:%M %p") if dt else str(raw).strip()


def fmt_resolution_date(raw):
    dt = _try_parse_dt(raw)
    return dt.strftime("%-d %b, %H:%M") if dt else str(raw).strip()


def incident_date_obj(raw):
    dt = _try_parse_dt(raw)
    return dt.date() if dt else None


def get_hyperlink(raw_row, col_map, field):
    idx = col_map.get(field)
    if idx is None or idx >= len(raw_row):
        return None
    cell = raw_row[idx]
    return cell.hyperlink.target if cell.hyperlink else None


def read_incidents(ws, tab_date):
    raw_rows = list(ws.iter_rows())
    rows = [[str(cell.value) if cell.value is not None else '' for cell in row]
            for row in raw_rows]
    if len(rows) < 2:
        return []

    col_map = build_col_map(rows[0])
    incidents = []

    for raw_row, row in zip(raw_rows[1:], rows[1:]):
        if not any(row):
            continue

        priority = get_field(row, col_map, 'priority')
        if priority not in ('P1', 'P2'):
            continue

        cause_raw = get_field(row, col_map, 'cause_type')
        cause_type = 'Internal' if 'internal' in cause_raw.lower() else 'External'

        status_raw = get_field(row, col_map, 'status').lower()
        if 'closed' in status_raw:
            status = 'Closed'
        elif 'ongoing' in status_raw or 'open' in status_raw:
            status = 'Open'
        else:
            status = 'TBC'

        jira_url = get_field(row, col_map, 'jira_url')
        jira_links = parse_jira_links(jira_url)
        jira_id = jira_links[0]['id'] if jira_links else ''
        jira_url = jira_links[0]['url'] if jira_links else ''

        pm_text = get_field(row, col_map, 'postmortem_url')
        pm_link = get_hyperlink(raw_row, col_map, 'postmortem_url')
        pm_url = pm_link or (pm_text if pm_text.startswith('http') else None)
        pm_label = pm_text if (pm_text and not pm_text.startswith('http')) else None

        date_raw = get_field(row, col_map, 'date')
        res_raw = get_field(row, col_map, 'resolution_dt') or get_field(row, col_map, 'resolution')

        date_obj = incident_date_obj(date_raw) or tab_date

        incidents.append({
            "title": get_field(row, col_map, 'title'),
            "date": fmt_incident_date(date_raw),
            "dateObj": date_obj.isoformat() if date_obj else None,
            "year": date_obj.year if date_obj else None,
            "monthNum": date_obj.month if date_obj else None,
            "monthKey": f"{date_obj.year:04d}-{date_obj.month:02d}" if date_obj else None,
            "causeType": cause_type,
            "status": status,
            "country": get_field(row, col_map, 'country'),
            "product": get_field(row, col_map, 'product'),
            "mttr": get_field(row, col_map, 'duration'),
            "resolution": fmt_resolution_date(res_raw),
            "impact": get_field(row, col_map, 'impact'),
            "cause": get_field(row, col_map, 'cause'),
            "jira": jira_id,
            "jiraUrl": jira_url,
            "jiras": jira_links,
            "postmortem": pm_label,
            "postmortemUrl": pm_url,
            "priority": priority,
        })

    return incidents


def find_last_internal(wb, months_back=6):
    tabs = get_monthly_sheets(wb, limit=months_back)
    best = None
    for tab_date, ws in tabs:
        for inc in read_incidents(ws, tab_date):
            if inc['causeType'] == 'Internal' and inc['dateObj']:
                d = datetime.date.fromisoformat(inc['dateObj'])
                if best is None or d > datetime.date.fromisoformat(best['date']):
                    best = {
                        "title": inc['title'],
                        "date": inc['dateObj'],
                        "jira": inc['jira'],
                        "jira_url": inc['jiraUrl'],
                    }
    return best


def build_payload():
    wb = download_workbook()
    tabs = [(d, ws) for d, ws in get_monthly_sheets(wb) if d >= MIN_DATE]

    out_incidents = []
    for d, ws in tabs:
        out_incidents.extend(read_incidents(ws, d))

    out_incidents.sort(key=lambda x: x['dateObj'] or '', reverse=True)

    months_present = sorted(set((i['year'], i['monthNum']) for i in out_incidents if i['monthNum']))
    month_meta = [
        {"key": f"{y:04d}-{m:02d}", "year": y, "num": m, "abbr": MONTH_ABBR[m], "full": f"{MONTH_FULL[m]} {y}"}
        for (y, m) in months_present
    ]

    last_internal = find_last_internal(wb)

    return {
        "generatedAt": datetime.date.today().isoformat(),
        "months": month_meta,
        "incidents": out_incidents,
        "lastInternal": last_internal,
    }


def main():
    print("Downloading spreadsheet …")
    payload = build_payload()
    print(f"Parsed {len(payload['incidents'])} incidents across {len(payload['months'])} months.")

    b64 = base64.b64encode(json.dumps(payload, separators=(',', ':')).encode()).decode()

    with open(TEMPLATE_PATH) as f:
        template = f.read()
    html = template.replace("__DATA_B64__", b64)

    with open(OUTPUT_PATH, "w") as f:
        f.write(html)

    print(f"Wrote {OUTPUT_PATH} ({len(html)} bytes)")


if __name__ == '__main__':
    main()
